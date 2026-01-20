#!/usr/bin/env python3
"""Generate mock CRM data for workshop tutorial"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

# Hebrew first names
first_names = [
    "יוסי", "דני", "מיכל", "רונית", "אבי", "שירה", "עומר", "נועה", "איתי", "מאיה",
    "גיל", "טל", "ליאור", "עדי", "רון", "דנה", "אורי", "יעל", "תומר", "שני",
    "אלון", "מור", "ניר", "הילה", "עידו", "רותם", "בועז", "ענת", "אייל", "לירון",
    "גלעד", "אפרת", "יניב", "קרן", "אסף", "נטע", "דוד", "סיגל", "משה", "אורלי",
    "חיים", "רינת", "יאיר", "גליה", "עמית", "מירב", "זיו", "ורד", "נדב", "שרון"
]

last_names = [
    "כהן", "לוי", "מזרחי", "פרץ", "ביטון", "אברהם", "דוד", "יוסף", "חיים", "מלכה",
    "אזולאי", "גבאי", "שמעון", "דהן", "אוחיון", "בן דוד", "אלון", "שרון", "ברק", "גולן",
    "רוזן", "פרידמן", "שפירא", "גולדברג", "ברנשטיין", "קליין", "וייס", "שוורץ", "הופמן", "זילברמן",
    "ממן", "חדד", "עמר", "סבג", "אבוטבול", "טל", "נחום", "אלמוג", "ירון", "שלום"
]

cities = ["תל אביב", "ירושלים", "חיפה", "באר שבע", "רמת גן", "פתח תקווה", "ראשון לציון", "נתניה", "אשדוד", "הרצליה"]

# Products and services
products = [
    {"name": "קורס AI למתחילים", "price": 1200, "type": "קורס"},
    {"name": "סדנת Claude Code", "price": 450, "type": "סדנה"},
    {"name": "קורס אוטומציה עסקית", "price": 1800, "type": "קורס"},
    {"name": "ייעוץ AI אישי - שעה", "price": 350, "type": "ייעוץ"},
    {"name": "חבילת ייעוץ - 5 שעות", "price": 1500, "type": "ייעוץ"},
    {"name": "סדנת Prompt Engineering", "price": 550, "type": "סדנה"},
    {"name": "קורס Python + AI", "price": 2200, "type": "קורס"},
    {"name": "וובינר חינמי - מבוא ל-AI", "price": 0, "type": "וובינר"},
    {"name": "סדנת אוטומציות ב-Make", "price": 400, "type": "סדנה"},
    {"name": "קורס מתקדם - AI Agents", "price": 2500, "type": "קורס"},
    {"name": "ייעוץ עסקי - חצי יום", "price": 2000, "type": "ייעוץ"},
    {"name": "סדנת ChatGPT למנהלים", "price": 600, "type": "סדנה"},
]

# Tags
tags_pool = [
    "VIP", "לקוח חוזר", "ממליץ", "מתעניין", "קר", "חם", "ליד חדש",
    "עסק קטן", "פרילנסר", "היי-טק", "סטארטאפ", "תאגיד", "עצמאי",
    "מתחיל", "מתקדם", "טכני", "לא טכני", "מנהל", "יזם", "מפתח",
    "השתתף בוובינר", "רכש קורס", "דורש מעקב", "פוטנציאל גבוה", "לא מעוניין"
]

statuses = ["פעיל", "לא פעיל", "VIP", "ליד", "בטיפול", "הושלם", "ממתין לתשובה", "בוגר"]

sources = ["אתר", "פייסבוק", "לינקדאין", "וואטסאפ", "המלצה", "גוגל", "אינסטגרם", "וובינר", "הרצאה"]

communication_types = ["שיחה", "אימייל", "וואטסאפ", "פגישה", "זום", "SMS"]

# Generate base date
today = datetime.now()

def random_date(start_days_ago, end_days_ago=0):
    days = random.randint(end_days_ago, start_days_ago)
    return today - timedelta(days=days)

def random_phone():
    prefix = random.choice(["050", "052", "053", "054", "058"])
    return f"{prefix}-{random.randint(1000000, 9999999)}"

def random_email(first, last):
    domain = random.choice(["gmail.com", "walla.co.il", "hotmail.com", "yahoo.com", "outlook.com"])
    return f"{first.lower()}.{last.lower()}@{domain}".replace(" ", "")

# Create workbook
wb = openpyxl.Workbook()

# Styles
header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
vip_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Generate 50 clients
clients = []
for i in range(50):
    client_id = f"C{str(i+1).zfill(4)}"
    first = random.choice(first_names)
    last = random.choice(last_names)

    client = {
        "id": client_id,
        "first_name": first,
        "last_name": last,
        "full_name": f"{first} {last}",
        "email": random_email(first, last),
        "phone": random_phone(),
        "city": random.choice(cities),
        "status": random.choice(statuses),
        "source": random.choice(sources),
        "tags": ", ".join(random.sample(tags_pool, k=random.randint(1, 4))),
        "created": random_date(365, 30),
        "last_contact": random_date(60, 0),
        "total_spent": 0,
        "notes": ""
    }
    clients.append(client)

# ============== SHEET 1: לקוחות ==============
ws1 = wb.active
ws1.title = "לקוחות"
ws1.sheet_view.rightToLeft = True

headers1 = ["מזהה", "שם פרטי", "שם משפחה", "אימייל", "טלפון", "עיר", "סטטוס", "מקור", "תגיות", "תאריך הצטרפות", "קשר אחרון", "הערות"]
ws1.append(headers1)
style_header(ws1)

for c in clients:
    ws1.append([
        c["id"], c["first_name"], c["last_name"], c["email"], c["phone"],
        c["city"], c["status"], c["source"], c["tags"],
        c["created"].strftime("%d/%m/%Y"), c["last_contact"].strftime("%d/%m/%Y"),
        ""
    ])

auto_width(ws1)

# ============== SHEET 2: היסטוריית רכישות ==============
ws2 = wb.create_sheet("היסטוריית רכישות")
ws2.sheet_view.rightToLeft = True

headers2 = ["מזהה עסקה", "מזהה לקוח", "שם לקוח", "מוצר", "סוג", "מחיר", "תאריך רכישה", "אמצעי תשלום", "סטטוס תשלום", "הערות"]
ws2.append(headers2)
style_header(ws2)

payment_methods = ["אשראי", "PayPal", "העברה בנקאית", "ביט", "מזומן"]
payment_statuses = ["שולם", "שולם", "שולם", "ממתין", "בוטל", "החזר"]

transaction_id = 1
for c in clients:
    num_purchases = random.choices([0, 1, 2, 3, 4], weights=[15, 35, 30, 15, 5])[0]
    for _ in range(num_purchases):
        product = random.choice(products)
        purchase_date = random_date(300, 5)
        payment_status = random.choices(payment_statuses, weights=[50, 20, 15, 8, 5, 2])[0]

        if payment_status == "שולם":
            c["total_spent"] += product["price"]

        notes = ""
        if random.random() < 0.2:
            notes = random.choice(["ביקש חשבונית", "קיבל הנחה 10%", "שילם בתשלומים", "הגיע דרך המלצה", "VIP"])

        ws2.append([
            f"T{str(transaction_id).zfill(5)}",
            c["id"],
            c["full_name"],
            product["name"],
            product["type"],
            product["price"],
            purchase_date.strftime("%d/%m/%Y"),
            random.choice(payment_methods),
            payment_status,
            notes
        ])
        transaction_id += 1

auto_width(ws2)

# ============== SHEET 3: תקשורת ==============
ws3 = wb.create_sheet("תקשורת")
ws3.sheet_view.rightToLeft = True

headers3 = ["מזהה", "מזהה לקוח", "שם לקוח", "סוג תקשורת", "תאריך", "נושא", "תקציר", "משימת המשך", "סטטוס"]
ws3.append(headers3)
style_header(ws3)

subjects = [
    "בירור על קורס", "מעקב אחרי רכישה", "שאלה טכנית", "בקשת הנחה", "תלונה",
    "בקשת מידע", "המלצה על קורס", "תיאום פגישה", "חידוש קשר", "מעקב ליד",
    "תודה על הקורס", "שאלה על תכנים", "בקשת תעודה", "עדכון פרטים", "ביטול"
]

summaries = [
    "הלקוח התעניין בקורס AI ורוצה פרטים נוספים",
    "שיחת מעקב - הלקוח מרוצה מהקורס",
    "בירור טכני על התקנת כלים",
    "הלקוח ביקש הנחה - הוסכם על 15%",
    "תלונה על איכות הצליל בהקלטה - נפתר",
    "שלחתי חומרים נוספים במייל",
    "הלקוח רוצה לשדרג לחבילת ייעוץ",
    "נקבעה פגישת ייעוץ ליום ראשון",
    "הלקוח לא זמין - לחזור בעוד שבוע",
    "ליד חם - מאוד מתעניין בקורס מתקדם",
    "הלקוח המליץ לחבר",
    "שלחתי קישור להרשמה",
    "הלקוח בתהליך החלטה",
    "עדכון מייל ומספר טלפון",
    "בוטלה הרשמה - החזר כספי בוצע"
]

follow_ups = [
    "", "", "", "לחזור בעוד שבוע", "לשלוח הצעת מחיר", "להזכיר על הסדנה",
    "לבדוק אם נרשם", "לשלוח סיכום פגישה", "", "לתאם שיחה נוספת"
]

comm_statuses = ["טופל", "טופל", "טופל", "פתוח", "ממתין לתגובה"]

comm_id = 1
for c in clients:
    num_comms = random.choices([0, 1, 2, 3, 4, 5], weights=[10, 25, 30, 20, 10, 5])[0]
    for _ in range(num_comms):
        comm_date = random_date(200, 1)
        ws3.append([
            f"CM{str(comm_id).zfill(5)}",
            c["id"],
            c["full_name"],
            random.choice(communication_types),
            comm_date.strftime("%d/%m/%Y %H:%M"),
            random.choice(subjects),
            random.choice(summaries),
            random.choice(follow_ups),
            random.choice(comm_statuses)
        ])
        comm_id += 1

auto_width(ws3)

# ============== SHEET 4: פרופיל מורחב ==============
ws4 = wb.create_sheet("פרופיל מורחב")
ws4.sheet_view.rightToLeft = True

headers4 = ["מזהה לקוח", "שם מלא", "גיל", "מקצוע", "חברה", "תפקיד", "רמת טכנית", "תחומי עניין", "סה״כ הוצאה", "ציון לקוח", "הערות פנימיות"]
ws4.append(headers4)
style_header(ws4)

professions = ["מהנדס תוכנה", "מנהל מוצר", "יזם", "מעצב", "מנכ״ל", "משווק", "מנהל פרויקטים", "עצמאי", "יועץ", "מפתח", "דאטה אנליסט", "HR", "מנהל כספים"]
companies = ["סטארטאפ", "היי-טק גדול", "עסק עצמאי", "חברת ייעוץ", "בנק", "קמעונאות", "חינוך", "בריאות", "ממשלה", "ללא"]
roles = ["מנהל", "עובד", "בעלים", "פרילנסר", "סטודנט", "מנהל בכיר", "צוות"]
tech_levels = ["מתחיל", "בסיסי", "בינוני", "מתקדם", "מומחה"]
interests = ["AI", "אוטומציה", "פיתוח", "עסקים", "שיווק", "ניהול", "פרודוקטיביות", "כלי No-Code", "Python", "נתונים"]

internal_notes = [
    "לקוח פוטנציאלי לחבילה עסקית",
    "מאוד אקטיבי בקהילה",
    "יש לו קהל - שותף פוטנציאלי",
    "צריך ליווי צמוד",
    "עצמאי - תקציב מוגבל",
    "מקבל החלטות מהר",
    "צריך זמן לחשוב",
    "רוצה לראות תוצאות מהירות",
    "מחפש פתרון מקיף",
    "לקוח VIP - יחס אישי",
    "",
    "",
    ""
]

for c in clients:
    age = random.randint(25, 60)
    score = random.randint(1, 10)
    if c["total_spent"] > 2000:
        score = min(10, score + 3)
    if "VIP" in c["tags"]:
        score = min(10, score + 2)

    ws4.append([
        c["id"],
        c["full_name"],
        age,
        random.choice(professions),
        random.choice(companies),
        random.choice(roles),
        random.choice(tech_levels),
        ", ".join(random.sample(interests, k=random.randint(2, 4))),
        c["total_spent"],
        score,
        random.choice(internal_notes)
    ])

auto_width(ws4)

# ============== SHEET 5: משימות ומעקב ==============
ws5 = wb.create_sheet("משימות ומעקב")
ws5.sheet_view.rightToLeft = True

headers5 = ["מזהה משימה", "מזהה לקוח", "שם לקוח", "סוג משימה", "תיאור", "תאריך יעד", "עדיפות", "סטטוס", "הוקצה ל", "הערות"]
ws5.append(headers5)
style_header(ws5)

task_types = ["מעקב", "שיחה", "הצעת מחיר", "תיאום", "תזכורת", "בירור", "משלוח חומר"]
priorities = ["נמוכה", "בינונית", "גבוהה", "דחוף"]
task_statuses = ["חדש", "בביצוע", "הושלם", "בהמתנה", "בוטל"]
assignees = ["אביץ", "צוות מכירות", "צוות תמיכה", ""]

task_descriptions = [
    "לחזור ללקוח עם הצעה",
    "לשלוח סילבוס של הקורס",
    "לתאם שיחת היכרות",
    "לשלוח קישור להרשמה",
    "מעקב אחרי רכישה",
    "לברר אם מעוניין בקורס הבא",
    "לשלוח תעודת סיום",
    "לבקש המלצה",
    "להציע שדרוג לחבילת VIP",
    "תזכורת על סדנה מחר"
]

task_id = 1
for c in random.sample(clients, 30):  # Tasks for 30 clients
    num_tasks = random.randint(1, 3)
    for _ in range(num_tasks):
        due_date = today + timedelta(days=random.randint(-10, 30))
        ws5.append([
            f"TK{str(task_id).zfill(4)}",
            c["id"],
            c["full_name"],
            random.choice(task_types),
            random.choice(task_descriptions),
            due_date.strftime("%d/%m/%Y"),
            random.choice(priorities),
            random.choice(task_statuses),
            random.choice(assignees),
            ""
        ])
        task_id += 1

auto_width(ws5)

# ============== SHEET 6: סטטיסטיקות ==============
ws6 = wb.create_sheet("סטטיסטיקות")
ws6.sheet_view.rightToLeft = True

# Summary stats
ws6.append(["סיכום נתונים"])
ws6["A1"].font = Font(bold=True, size=14)
ws6.append([])
ws6.append(["סה״כ לקוחות", len(clients)])
ws6.append(["לקוחות פעילים", sum(1 for c in clients if c["status"] == "פעיל")])
ws6.append(["לקוחות VIP", sum(1 for c in clients if c["status"] == "VIP")])
ws6.append(["סה״כ הכנסות", f"₪{sum(c['total_spent'] for c in clients):,}"])
ws6.append([])
ws6.append(["התפלגות לפי מקור"])
for source in set(c["source"] for c in clients):
    count = sum(1 for c in clients if c["source"] == source)
    ws6.append([source, count])

ws6.append([])
ws6.append(["התפלגות לפי עיר"])
for city in set(c["city"] for c in clients):
    count = sum(1 for c in clients if c["city"] == city)
    ws6.append([city, count])

auto_width(ws6)

# Save
output_path = "/Users/aviz/architect-workshops/workshops/2026-01-20-claudosh-spaceship/tutorial/mock_crm_data.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Sheets: {wb.sheetnames}")
